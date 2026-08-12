import logging

from django.db.models import Q, Sum, Value, DecimalField
from django.db.models.functions import Coalesce
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from expenses.models import ExpenseRequest
from retirement.models import RetirementForm
from core.models import Department
import io
import calendar
from django.utils import timezone
from django.conf import settings

logger = logging.getLogger(__name__)


def _default_date_range():
    today = timezone.localdate()
    start = today.replace(day=1).isoformat()
    end = today.replace(day=calendar.monthrange(today.year, today.month)[1]).isoformat()
    return start, end


@login_required
def reports_dashboard(request):
    return redirect('reports:expenses')


@login_required
def expenses_report(request):
    try:
        user = request.user
        is_admin = user.is_staff or user.is_superuser
        is_treasurer = hasattr(user, 'treasurer_profile')
        is_approver = hasattr(user, 'approver_profile')
        approver = None
        if is_admin or is_treasurer:
            qs = ExpenseRequest.objects.all()
        elif is_approver:
            approver = user.approver_profile
            if approver.level == 'first':
                qs = ExpenseRequest.objects.filter(department__in=approver.departments.all())
            else:
                qs = ExpenseRequest.objects.all()
        else:
            qs = ExpenseRequest.objects.filter(submitted_by=user)

        status_filter = request.GET.get('status', '').strip()
        search = request.GET.get('search', '').strip()
        date_from = request.GET.get('date_from', '').strip() or _default_date_range()[0]
        date_to = request.GET.get('date_to', '').strip() or _default_date_range()[1]
        payment_filter = request.GET.get('payment', '').strip()
        department_id = request.GET.get('department', '').strip()

        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if search:
            qs = qs.filter(
                Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(form_number__icontains=search)
                | Q(reason__icontains=search)
            )
        if payment_filter == 'paid':
            qs = qs.filter(is_paid=True)
        elif payment_filter == 'unpaid':
            qs = qs.filter(is_paid=False)
        if department_id:
            qs = qs.filter(department_id=department_id)

        qs = qs.select_related('department', 'submitted_by')
        status_choices = ExpenseRequest.STATUS_CHOICES
        approved_count = qs.filter(status__in=['approved', 'paid']).count()
        pending_count = qs.exclude(status__in=['approved', 'paid']).count()
        expense_total = qs.aggregate(total=Coalesce(Sum('total_amount'), Value(0), output_field=DecimalField()))['total']

        departments = Department.objects.all()
        if is_approver and not is_admin and not is_treasurer and approver and approver.level == 'first':
            departments = approver.departments.all()

        return render(request, 'reports/expenses.html', {
        'expenses': list(qs),
        'is_admin': is_admin,
        'is_treasurer': is_treasurer,
        'is_approver': is_approver,
        'user': request.user,
        'departments': departments,
        'status_choices': status_choices,
        'status_filter': status_filter,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'payment_filter': payment_filter,
        'department_id': department_id,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'expense_total': expense_total,
    })
    except Exception as exc:
        logger.exception('Unexpected error rendering expenses report dashboard: %s', exc)
        messages.error(request, 'An error occurred while loading the expenses report. Please contact support.')
        return render(request, 'reports/expenses.html', {
            'expenses': [],
            'is_admin': False,
            'is_treasurer': False,
            'is_approver': False,
            'user': request.user,
            'departments': Department.objects.none(),
            'status_choices': ExpenseRequest.STATUS_CHOICES,
            'status_filter': '',
            'search': '',
            'date_from': _default_date_range()[0],
            'date_to': _default_date_range()[1],
            'payment_filter': '',
            'department_id': '',
            'approved_count': 0,
            'pending_count': 0,
            'expense_total': 0,
        })


@login_required
def retirement_report(request):
    user = request.user
    is_admin = user.is_staff or user.is_superuser
    is_treasurer = hasattr(user, 'treasurer_profile')
    is_approver = hasattr(user, 'approver_profile')
    approver = None
    if is_admin or is_treasurer:
        qs = RetirementForm.objects.all()
    elif is_approver:
        approver = user.approver_profile
        if approver.level == 'first':
            qs = RetirementForm.objects.filter(department__in=approver.departments.all())
        else:
            qs = RetirementForm.objects.all()
    else:
        qs = RetirementForm.objects.filter(submitted_by=user)

    try:
        status_filter = request.GET.get('status', '').strip()
        search = request.GET.get('search', '').strip()
        date_from = request.GET.get('date_from', '').strip() or _default_date_range()[0]
        date_to = request.GET.get('date_to', '').strip() or _default_date_range()[1]
        department_id = request.GET.get('department', '').strip()

        if date_from:
            qs = qs.filter(date_of_request__gte=date_from)
        if date_to:
            qs = qs.filter(date_of_request__lte=date_to)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if search:
            qs = qs.filter(
                Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(form_number__icontains=search)
                | Q(reason__icontains=search)
                | Q(exp_request_form_no__icontains=search)
            )
        if department_id:
            qs = qs.filter(department_id=department_id)

        qs = qs.select_related('department', 'submitted_by')
        status_choices = RetirementForm.STATUS_CHOICES
        approved_count = qs.filter(status__in=['approved', 'paid']).count()
        pending_count = qs.exclude(status__in=['approved', 'paid']).count()
        retirement_total = qs.aggregate(total=Coalesce(Sum('remaining_amount'), Value(0), output_field=DecimalField()))['total']

        departments = Department.objects.all()
        if is_approver and not is_admin and not is_treasurer and approver and approver.level == 'first':
            departments = approver.departments.all()

        return render(request, 'reports/retirement.html', {
        'retirements': list(qs),
        'is_admin': is_admin,
        'is_treasurer': is_treasurer,
        'is_approver': is_approver,
        'user': request.user,
        'departments': departments,
        'status_choices': status_choices,
        'status_filter': status_filter,
        'search': search,
        'date_from': date_from,
        'date_to': date_to,
        'department_id': department_id,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'retirement_total': retirement_total,
    })
    except Exception as exc:
        logger.exception('Unexpected error rendering retirement report dashboard: %s', exc)
        messages.error(request, 'An error occurred while loading the retirement report. Please contact support.')
        return render(request, 'reports/retirement.html', {
            'retirements': [],
            'is_admin': False,
            'is_treasurer': False,
            'is_approver': False,
            'user': request.user,
            'departments': Department.objects.none(),
            'status_choices': RetirementForm.STATUS_CHOICES,
            'status_filter': '',
            'search': '',
            'date_from': _default_date_range()[0],
            'date_to': _default_date_range()[1],
            'approved_count': 0,
            'pending_count': 0,
            'retirement_total': 0,
        })


@login_required
def download_expense_report(request):
    if getattr(settings, 'DISABLE_EXCEL_DOWNLOAD', False):
        return HttpResponse('Expense Excel downloads are disabled on this deployment.', status=503)
    user = request.user
    is_admin = user.is_staff or user.is_superuser
    is_treasurer = hasattr(user, 'treasurer_profile')
    is_approver = hasattr(user, 'approver_profile')

    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if not request.GET:
        date_from, date_to = _default_date_range()

    if is_admin or is_treasurer:
        qs = ExpenseRequest.objects.all()
    elif is_approver:
        approver = user.approver_profile
        if approver.level == 'first':
            qs = ExpenseRequest.objects.filter(department__in=approver.departments.all())
        else:
            qs = ExpenseRequest.objects.all()
    else:
        qs = ExpenseRequest.objects.filter(submitted_by=user)
    status_filter = request.GET.get('status', '').strip()
    search = request.GET.get('search', '').strip()
    payment_filter = request.GET.get('payment', '').strip()
    department_id = request.GET.get('department', '').strip()

    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(form_number__icontains=search)
            | Q(reason__icontains=search)
        )
    if payment_filter == 'paid':
        qs = qs.filter(is_paid=True)
    elif payment_filter == 'unpaid':
        qs = qs.filter(is_paid=False)
    if department_id:
        qs = qs.filter(department_id=department_id)

    # Generate Excel for all users
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse('openpyxl not available', status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Expense Report'
    headers = ['#', 'Form No', 'Name', 'Department', 'Date', 'Reason', 'Amount (TZS)', 'Status', 'Paid']
    ws.append(headers)

    total = 0
    for i, e in enumerate(qs, start=1):
        amt = float(getattr(e, 'total_amount', 0) or 0)
        ws.append([
            i,
            getattr(e, 'form_number', ''),
            f"{getattr(e,'first_name','')} {getattr(e,'last_name','')}",
            str(getattr(e, 'department', '') or ''),
            str(getattr(e, 'date', '')),
            getattr(e, 'reason', ''),
            amt,
            getattr(e, 'get_status_display', lambda: '')(),
            'Yes' if getattr(e, 'is_paid', False) else 'No',
        ])
        total += amt

    ws.append(['', '', '', '', '', 'TOTAL', float(total), '', ''])
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col:
            v = '' if cell.value is None else str(cell.value)
            if v:
                max_length = max(max_length, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_length + 2)

    buffer_xl = io.BytesIO()
    wb.save(buffer_xl)
    buffer_xl.seek(0)
    response = HttpResponse(buffer_xl.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="expense_report.xlsx"'
    return response


@login_required
def download_retirement_report(request):
    if getattr(settings, 'DISABLE_EXCEL_DOWNLOAD', False):
        return HttpResponse('Retirement Excel downloads are disabled on this deployment.', status=503)
    user = request.user
    is_admin = user.is_staff or user.is_superuser
    is_treasurer = hasattr(user, 'treasurer_profile')
    is_approver = hasattr(user, 'approver_profile')

    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    if not request.GET:
        date_from, date_to = _default_date_range()

    if is_admin or is_treasurer:
        qs = RetirementForm.objects.all()
    elif is_approver:
        approver = user.approver_profile
        if approver.level == 'first':
            qs = RetirementForm.objects.filter(department__in=approver.departments.all())
        else:
            qs = RetirementForm.objects.all()
    else:
        qs = RetirementForm.objects.filter(submitted_by=user)
    status_filter = request.GET.get('status', '').strip()
    search = request.GET.get('search', '').strip()
    department_id = request.GET.get('department', '').strip()

    if date_from:
        qs = qs.filter(date_of_request__gte=date_from)
    if date_to:
        qs = qs.filter(date_of_request__lte=date_to)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if search:
        qs = qs.filter(
            Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(form_number__icontains=search)
            | Q(reason__icontains=search)
            | Q(exp_request_form_no__icontains=search)
        )
    if department_id:
        qs = qs.filter(department_id=department_id)

    # Generate Excel for all users
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse('openpyxl not available', status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Retirement Report'
    headers = ['#', 'Form No', 'Exp Req No', 'Name', 'Department', 'Date Req.', 'Date Ret.', 'Remaining', 'Status']
    ws.append(headers)

    total = 0
    for i, r in enumerate(qs, start=1):
        rem = float(getattr(r, 'remaining_amount', 0) or 0)
        ws.append([
            i,
            getattr(r, 'form_number', ''),
            getattr(r, 'exp_request_form_no', '') or '',
            f"{getattr(r,'first_name','')} {getattr(r,'last_name','')}",
            str(getattr(r, 'department', '') or ''),
            str(getattr(r, 'date_of_request', '')),
            str(getattr(r, 'date_of_retirement', '')),
            rem,
            getattr(r, 'get_status_display', lambda: '')(),
        ])
        total += rem

    ws.append(['', '', '', '', '', '', 'TOTAL', float(total), ''])
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in col:
            v = '' if cell.value is None else str(cell.value)
            if v:
                max_length = max(max_length, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, max_length + 2)

    buffer_xl = io.BytesIO()
    wb.save(buffer_xl)
    buffer_xl.seek(0)
    response = HttpResponse(buffer_xl.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="retirement_report.xlsx"'
    return response
